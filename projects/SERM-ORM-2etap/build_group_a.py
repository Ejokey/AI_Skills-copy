# -*- coding: utf-8 -*-
import openpyxl
from copy import copy

REF = 'reference/Образец_Эн+_En+_Group_смета.xlsx'
NIT = 'smetas/1ISadLOK_Ново-Иркутская ТЭЦ_смета.xlsx'
OUT = 'Сметы_Группа_А_v2.xlsx'

wb = openpyxl.Workbook()
wb.remove(wb.active)


def copy_sheet_verbatim(src_path, src_sheet_name, dst_name):
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb[src_sheet_name]
    dst_ws = wb.create_sheet(dst_name)

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = cell.number_format

    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    return dst_ws


copy_sheet_verbatim(REF, 'Лист1', 'Эн+ (En+ Group)')
copy_sheet_verbatim(NIT, 'Лист1', 'Ново-Иркутская ТЭЦ')

ref_wb = openpyxl.load_workbook(REF)
ref_ws = ref_wb['Лист1']
ref_col_widths = {c: d.width for c, d in ref_ws.column_dimensions.items()}


def build_fresh_sheet(name, m, c7_text, d7_base):
    ws = wb.create_sheet(name)

    for col_letter, width in ref_col_widths.items():
        ws.column_dimensions[col_letter].width = width

    def f(base):
        return f'={base}*{m}' if m != 1 else f'={base}'

    values = {
        'A1': 'Месяц работы по объекту',
        'B1': 'Оказываемый блок услуг',
        'C1': 'Перечень услуг и их количество (при необходимости пояснения)',
        'D1': 'Стоимость блока услуг',

        'A2': 1,
        'B2': 'Первичный аудит поисковой выдачи',
        'C2': 'Сбор и анализ упоминаний объектов в соцсетях и СМИ, анализ поисковой выдачи и ИИ-подсказок. Определение тональности упоминаний, структуры негатива, проблемных площадок и анализ рисков.',
        'D2': f(22000),

        'B3': 'Разработка стратегии работ на год',
        'C3': 'Разработка стратегии работы с информационным полем',
        'D3': f(31000),

        'B4': 'Первичный мониторинг ИИ-подсказок и генеративной выдачи',
        'C4': 'Формирование запросов, регулярная проверка, выявление искажений, источники влияния, рекомендации, динамика',
        'D4': f(11000),

        'B5': 'Итого',
        'D5': '=SUM(D2:D4)',

        'A6': 'Начиная с 2 месяца',
        'B6': 'Регулярный аудит поисковой выдачи, ИИ-подсказок и генеративной выдачи',
        'C6': 'Ежемесячно',
        'D6': f(95000),

        'B7': 'ORM: подготовка комментариев для поддержки положительной и нейтральной информации о компании на сайтах отзывов, карьерных ресурсах и картографических сервисах',
        'C7': c7_text,
        'D7': f(d7_base),

        'B8': 'SERM в соответствии с п.3.5 ТЗ',
        'C8': 'Регулярный анализ поисковой выдачи, отслеживание изменения позиций материалов в поисковой выдаче, отчетность (в т.ч. менеджмент проекта) (продвижение не менее 20 площадок)',
        'D8': f(114720),

        'C9': 'Корректировка поисковых подсказок',
        'D9': f(24000),

        'C10': 'Корректировка ИИ-блока поисковой выдачи',
        'D10': f(42000),

        'C11': 'Подготовка текстов (SEO-статей об объектах) для размещения на внешних площадках',
        'D11': f(14000),

        'B12': 'Итого в месяц',
        'D12': '=SUM(D6:D11)',
    }

    for r in range(1, 13):
        for col in 'ABCD':
            coord = f'{col}{r}'
            src_style_cell = ref_ws[coord]
            dst_cell = ws[coord]
            dst_cell.value = values.get(coord)
            dst_cell.font = copy(src_style_cell.font)
            dst_cell.border = copy(src_style_cell.border)
            dst_cell.fill = copy(src_style_cell.fill)
            dst_cell.alignment = copy(src_style_cell.alignment)
            dst_cell.number_format = 'General'

    ws.merge_cells('A2:A5')
    ws.merge_cells('A6:A12')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B8:B11')
    ws.merge_cells('B12:C12')

    return ws


build_fresh_sheet('РУСАЛ', m=1, c7_text='До 10 отзывов и до 50 комментариев к отзывам', d7_base=70000)
build_fresh_sheet('ТЭЦ-9 + Ангарская ТЭЦ-9', m=1, c7_text='До 5 отзывов и до 10 комментариев к отзывам', d7_base=20000)
build_fresh_sheet('ИркАЗ', m=2, c7_text='До 5 отзывов и до 10 комментариев к отзывам', d7_base=20000)
build_fresh_sheet('САЗ', m=2, c7_text='До 5 отзывов и до 10 комментариев к отзывам', d7_base=20000)
build_fresh_sheet('АГК', m=2, c7_text='До 5 отзывов и до 10 комментариев к отзывам', d7_base=20000)

print('BEFORE reorder/save: bold A1 NIT =', wb['Ново-Иркутская ТЭЦ']['A1'].font.bold, 'border=', wb['Ново-Иркутская ТЭЦ']['A1'].border.left.style)
print('BEFORE reorder/save: bold A1 Эн+ =', wb['Эн+ (En+ Group)']['A1'].font.bold, 'border=', wb['Эн+ (En+ Group)']['A1'].border.left.style)
print('BEFORE reorder/save: bold A1 РУСАЛ =', wb['РУСАЛ']['A1'].font.bold, 'border=', wb['РУСАЛ']['A1'].border.left.style)

order = ['Эн+ (En+ Group)', 'РУСАЛ', 'Ново-Иркутская ТЭЦ', 'ТЭЦ-9 + Ангарская ТЭЦ-9', 'ИркАЗ', 'САЗ', 'АГК']
wb._sheets = [wb[name] for name in order]

wb.save(OUT)
print('saved', OUT, wb.sheetnames)

wb2 = openpyxl.load_workbook(OUT)
print('AFTER reload: bold A1 NIT =', wb2['Ново-Иркутская ТЭЦ']['A1'].font.bold, 'border=', wb2['Ново-Иркутская ТЭЦ']['A1'].border.left.style)
print('AFTER reload: bold A1 Эн+ =', wb2['Эн+ (En+ Group)']['A1'].font.bold, 'border=', wb2['Эн+ (En+ Group)']['A1'].border.left.style)
print('AFTER reload: bold A1 РУСАЛ =', wb2['РУСАЛ']['A1'].font.bold, 'border=', wb2['РУСАЛ']['A1'].border.left.style)
