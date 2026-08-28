# -*- coding: utf-8 -*-
import openpyxl
import re

PATH = 'Сметы_Группа_А.xlsx'
RUB_FMT = '#,##0.00\\ "₽"'

wb = openpyxl.load_workbook(PATH)

for name in wb.sheetnames:
    ws = wb[name]

    def resolve(coord):
        cell = ws[coord]
        v = cell.value
        if isinstance(v, str) and v.startswith('='):
            expr = v[1:]
            m = re.match(r'^(\d+)\*(\d+)$', expr)
            if m:
                return int(m.group(1)) * int(m.group(2))
            m = re.match(r'^(\d+)$', expr)
            if m:
                return int(expr)
            m = re.match(r'^SUM\(D(\d+):D(\d+)\)$', expr)
            if m:
                lo, hi = int(m.group(1)), int(m.group(2))
                return sum(resolve(f'D{r}') for r in range(lo, hi + 1))
            raise ValueError(f'{name}!{coord}: unsupported formula {v}')
        return v

    for r in range(2, 13):
        coord = f'D{r}'
        if ws[coord].value is None:
            continue
        val = resolve(coord)
        ws[coord].value = val
        ws[coord].number_format = RUB_FMT

wb.save(PATH)
print('done')
