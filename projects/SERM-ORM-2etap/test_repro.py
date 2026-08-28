# -*- coding: utf-8 -*-
import openpyxl
from copy import copy

REF = 'reference/Образец_Эн+_En+_Group_смета.xlsx'
NIT = 'smetas/1ISadLOK_Ново-Иркутская ТЭЦ_смета.xlsx'
OUT = '/tmp_repro.xlsx'

wb = openpyxl.Workbook()
wb.remove(wb.active)


def copy_sheet_verbatim(src_path, src_sheet_name, dst_name):
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb[src_sheet_name]
    dst_ws = wb.create_sheet(dst_name)

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = cell.number_format

    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    return dst_ws


copy_sheet_verbatim(REF, 'Лист1', 'Эн+ (En+ Group)')
copy_sheet_verbatim(NIT, 'Лист1', 'Ново-Иркутская ТЭЦ')

ref_wb = openpyxl.load_workbook(REF)
ref_ws = ref_wb['Лист1']
ref_col_widths = {c: d.width for c, d in ref_ws.column_dimensions.items()}


def build_fresh_sheet(name, m, c7_text, d7_base):
    ws = wb.create_sheet(name)
    for col_letter, width in ref_col_widths.items():
        ws.column_dimensions[col_letter].width = width

    def f(base):
        return f'={base}*{m}' if m != 1 else f'={base}'

    values = {
        'A1': 'x', 'D2': f(22000), 'D5': '=SUM(D2:D4)', 'D12': '=SUM(D6:D11)',
    }
    for r in range(1, 13):
        for col in 'ABCD':
            coord = f'{col}{r}'
            src_style_cell = ref_ws[coord]
            dst_cell = ws[coord]
            dst_cell.value = values.get(coord)
            dst_cell.font = copy(src_style_cell.font)
            dst_cell.border = copy(src_style_cell.border)
            dst_cell.fill = copy(src_style_cell.fill)
            dst_cell.alignment = copy(src_style_cell.alignment)
            dst_cell.number_format = 'General'
    ws.merge_cells('A2:A5')
    ws.merge_cells('A6:A12')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B8:B11')
    ws.merge_cells('B12:C12')
    return ws


build_fresh_sheet('РУСАЛ', m=1, c7_text='x', d7_base=70000)

print('BEFORE reorder: bold A1 NIT =', wb['Ново-Иркутская ТЭЦ']['A1'].font.bold)

order = ['Эн+ (En+ Group)', 'РУСАЛ', 'Ново-Иркутская ТЭЦ']
wb._sheets = [wb[name] for name in order]

print('AFTER reorder (in-memory): bold A1 NIT =', wb['Ново-Иркутская ТЭЦ']['A1'].font.bold)

wb.save(OUT)
wb2 = openpyxl.load_workbook(OUT)
print('AFTER save+reload: bold A1 NIT =', wb2['Ново-Иркутская ТЭЦ']['A1'].font.bold)
print('AFTER save+reload: border A1 NIT =', wb2['Ново-Иркутская ТЭЦ']['A1'].border.left.style)
